"""
Excel formatting utilities for the URL Migration Tool.

This module provides functions to apply conditional formatting to Excel files
exported by the URL Migration Tool, such as highlighting matching URLs in green.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from typing import List, Optional
import logging






def highlight_matching_urls(excel_path: Path, logger: Optional[logging.Logger] = None) -> None:
    """
    Apply conditional formatting to highlight URLs that match the "Best redirect" URL in green.
    
    Args:
        excel_path (Path): Path to the Excel file to format
        logger (Optional[logging.Logger]): Logger instance for logging messages
    """
    try:
        # Load the workbook
        workbook = load_workbook(excel_path)
        
        # Get the mapping sheet
        mapping_sheet = workbook['Mapping']
        
        # Find the column indices for algorithm columns and "Best redirect" column
        header_row = 1
        algorithm_columns = []
        best_redirect_col = None
        score_columns = []
        
        # Get all column headers
        headers = []
        for col_idx, cell in enumerate(mapping_sheet[header_row], 1):
            if cell.value and isinstance(cell.value, str):
                headers.append((col_idx, cell.value))
                
                # Find the "Best redirect" column
                if cell.value == "Best redirect":
                    best_redirect_col = col_idx
                # Identify score columns (skip them)
                elif cell.value.endswith('_score'):
                    score_columns.append(col_idx)
        
        # Identify algorithm columns (all columns except scores and specific non-URL columns)
        excluded_keywords = ["score", "agreement", "totscore", "best redirect", "status", "count"]
        for col_idx, header in headers:
            # Skip score columns and other non-URL columns
            if (col_idx not in score_columns and 
                not any(keyword in header.lower() for keyword in excluded_keywords) and
                "url_404" not in header.lower()):  # Skip the 404 URL column
                algorithm_columns.append(col_idx)
        
        if not algorithm_columns or best_redirect_col is None:
            if logger:
                logger.warning(f"Could not find algorithm columns or 'Best redirect' column in {excel_path}")
                logger.info(f"Headers found: {[h for _, h in headers]}")
                logger.info(f"Algorithm columns detected: {algorithm_columns}")
            return
        
        # Define green fill
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # Apply conditional formatting
        for row_idx in range(2, mapping_sheet.max_row + 1):
            # Get the "Best redirect" URL for this row
            best_redirect_cell = mapping_sheet.cell(row=row_idx, column=best_redirect_col)
            best_redirect_url = best_redirect_cell.value
            
            if best_redirect_url:
                # Check each algorithm column and apply green fill if it matches
                for col_idx in algorithm_columns:
                    cell = mapping_sheet.cell(row=row_idx, column=col_idx)
                    if cell.value and str(cell.value).strip() == str(best_redirect_url).strip():
                        cell.fill = green_fill
        
        # Save the workbook
        workbook.save(excel_path)
        
        if logger:
            logger.info(f"Applied conditional formatting to {excel_path}")
            logger.info(f"Highlighted {len(algorithm_columns)} algorithm columns with matching URLs")
            
    except Exception as e:
        if logger:
            logger.error(f"Error applying conditional formatting: {str(e)}")
        else:
            print(f"Error applying conditional formatting: {str(e)}")



















class ScoreHighlighter:
    """
    Class to highlight score columns in Excel files using a color gradient
    from red (low scores) to green (high scores).
    """
    
    def __init__(self, excel_path: Path, logger: Optional[logging.Logger] = None):
        """
        Initialize the ScoreHighlighter with the path to the Excel file.
        
        Args:
            excel_path (Path): Path to the Excel file to format
            logger (Optional[logging.Logger]): Logger instance for logging messages
        """
        self.excel_path = excel_path
        self.logger = logger
        self.workbook = None
        self.mapping_sheet = None
    
    def load_workbook(self):
        """Load the Excel workbook and get the mapping sheet."""
        try:
            self.workbook = load_workbook(self.excel_path)
            self.mapping_sheet = self.workbook['Mapping']
            return True
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error loading workbook: {str(e)}")
            return False
    
    def identify_score_columns(self):
        """
        Identify columns containing score values (with "_score" in column name).
        
        Returns:
            List[int]: List of column indices for score columns
        """
        score_columns = []
        header_row = 1
        
        for col_idx, cell in enumerate(self.mapping_sheet[header_row], 1):
            if cell.value and isinstance(cell.value, str) and "_score" in cell.value.lower():
                score_columns.append(col_idx)
        
        if self.logger:
            self.logger.info(f"Identified {len(score_columns)} score columns")
        
        return score_columns
    
    def get_color_for_score(self, score):
        """
        Get a color fill based on the score value (red to yellow to green gradient).
        
        Args:
            score (float): Score value between 0 and 1
            
        Returns:
            PatternFill: OpenPyXL PatternFill object with appropriate color
        """
        if score is None or not isinstance(score, (int, float)):
            return PatternFill(fill_type=None)
        
        # Ensure score is between 0 and 1
        score = max(0, min(1, float(score)))
        
        # Red component (decreases from 255 to 0 as score goes from 0.5 to 1)
        red = int(255 if score < 0.5 else 255 * 2 * (1 - score))
        
        # Green component (increases from 0 to 255 as score goes from 0 to 0.5)
        green = int(255 if score > 0.5 else 255 * 2 * score)
        
        # Blue component (keep low for better contrast)
        blue = 0
        
        # Convert to hex color code
        color_code = f"{red:02x}{green:02x}{blue:02x}"
        
        return PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
    
    def apply_gradient_to_score_columns(self, score_columns):
        """
        Apply color gradient to score columns based on their values.
        
        Args:
            score_columns (List[int]): List of column indices for score columns
        """
        # Start from row 2 (after header)
        for row_idx in range(2, self.mapping_sheet.max_row + 1):
            for col_idx in score_columns:
                cell = self.mapping_sheet.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None:
                        score = float(cell.value)
                        cell.fill = self.get_color_for_score(score)
                except (ValueError, TypeError):
                    # Skip cells that can't be converted to float
                    continue
    
    def highlight_scores(self):
        """
        Process the Excel file and highlight score columns with a color gradient.
        """
        if not self.load_workbook():
            return
        
        score_columns = self.identify_score_columns()
        
        if not score_columns:
            if self.logger:
                self.logger.warning("No score columns found in the Excel file")
            return
        
        self.apply_gradient_to_score_columns(score_columns)
        
        try:
            self.workbook.save(self.excel_path)
            if self.logger:
                self.logger.info(f"Applied score highlighting to {len(score_columns)} columns in {self.excel_path}")
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error saving workbook after highlighting: {str(e)}")

def highlight_score_columns(excel_path: Path, logger: Optional[logging.Logger] = None) -> None:
    """
    Apply color gradient formatting to score columns in the Excel file.
    
    Args:
        excel_path (Path): Path to the Excel file to format
        logger (Optional[logging.Logger]): Logger instance for logging messages
    """
    try:
        highlighter = ScoreHighlighter(excel_path, logger)
        highlighter.highlight_scores()
    except Exception as e:
        if logger:
            logger.error(f"Error applying score highlighting: {str(e)}")
        else:
            print(f"Error applying score highlighting: {str(e)}")